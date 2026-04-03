#!/usr/bin/env python3
"""
检查并修正绩效表中「自评分低于上级打分」的不一致（与研发中心季度绩效考评表-专业线布局一致）。

业务区（任务行，D 列有任务描述）：
  I=自评完成时间, J=自评完成质量, K=完成时间直接上级, L=完成质量直接上级, M=完成质量隔级上级

综合素质区（「创新能力」至「遵章守纪」或「工作计划」）：
  K=权重, L=自评, M=直接上级, N=隔级上级

用法：
  python check_self_and_leader_score.py check <目录>    # 检查并写入结果文件
  python check_self_and_leader_score.py modify          # 按结果文件修正（默认读当前目录下结果文件）
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import openpyxl

RESULT_FILENAME = "check_self_and_leader_score_result.txt"

# 业务区列（1-based）
COL_SELF_TIME = 9   # I
COL_SELF_QUALITY = 10  # J
COL_LEADER_TIME = 11  # K
COL_LEADER_QUALITY_DIRECT = 12  # L
COL_LEADER_QUALITY_SKIP = 13  # M

# 综合素质区
COL_PERSONAL_SELF = 12  # L
COL_PERSONAL_DIRECT = 13  # M
COL_PERSONAL_SKIP = 14  # N


def iter_xlsx_files(directory: str) -> list[Path]:
    root = Path(directory).resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"不是目录: {root}")
    out: list[Path] = []
    for dirpath, _dirnames, filenames in os.walk(root):
        for name in filenames:
            if name.startswith("."):
                continue
            if not name.endswith(".xlsx"):
                continue
            if name.startswith("~$") or name.startswith(".~"):
                continue
            out.append(Path(dirpath) / name)
    return sorted(out)


def to_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def task_section_bounds(ws) -> tuple[int, int] | None:
    if ws["D7"].value != "任务目标":
        return None
    end_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "业务完成综合分（项目平均分）":
            end_row = r - 1
            break
    if end_row is None or end_row < 9:
        return None
    return 9, end_row


def personal_section_bounds(ws) -> tuple[int, int] | None:
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "评估维度":
            header_row = r
            break
    if header_row is None:
        return None

    start_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "创新能力":
            start_row = r
            break
    if start_row is None:
        return None

    end_row = None
    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a == "遵章守纪":
            end_row = r
            break
        if a == "工作计划":
            end_row = r
            break
    if end_row is None:
        return None
    return start_row, end_row


def sheet_name_from_c2(ws) -> str:
    v = ws["C2"].value
    if v is None:
        return ""
    return str(v).strip()


def check_sheet(ws, display_name: str) -> list[str]:
    """返回该工作表上的问题描述行（不含姓名前缀）。"""
    issues: list[str] = []

    bounds = task_section_bounds(ws)
    if bounds:
        first_r, last_r = bounds
        for r in range(first_r, last_r + 1):
            if not ws.cell(row=r, column=4).value:
                continue
            st = to_float(ws.cell(row=r, column=COL_SELF_TIME).value)
            sq = to_float(ws.cell(row=r, column=COL_SELF_QUALITY).value)
            lt = to_float(ws.cell(row=r, column=COL_LEADER_TIME).value)
            lq = to_float(ws.cell(row=r, column=COL_LEADER_QUALITY_DIRECT).value)
            mq = to_float(ws.cell(row=r, column=COL_LEADER_QUALITY_SKIP).value)

            if st is not None and lt is not None and st < lt:
                issues.append(
                    f"业务第{r}行，完成时间自评（{st:g}）低于直接上级打分（{lt:g}）"
                )
            if sq is not None:
                if lq is not None and sq < lq:
                    issues.append(
                        f"业务第{r}行，完成质量自评（{sq:g}）低于直接上级打分（{lq:g}）"
                    )
                if mq is not None and sq < mq:
                    issues.append(
                        f"业务第{r}行，完成质量自评（{sq:g}）低于隔级上级打分（{mq:g}）"
                    )

    pb = personal_section_bounds(ws)
    if pb:
        p_first, p_last = pb
        for r in range(p_first, p_last + 1):
            sv = to_float(ws.cell(row=r, column=COL_PERSONAL_SELF).value)
            md = to_float(ws.cell(row=r, column=COL_PERSONAL_DIRECT).value)
            sk = to_float(ws.cell(row=r, column=COL_PERSONAL_SKIP).value)
            dim = ws.cell(row=r, column=1).value
            dim_s = str(dim).strip() if dim else f"第{r}行"

            if sv is not None:
                if md is not None and sv < md:
                    issues.append(
                        f"综合素质「{dim_s}」，自评（{sv:g}）低于直接上级打分（{md:g}）"
                    )
                if sk is not None and sv < sk:
                    issues.append(
                        f"综合素质「{dim_s}」，自评（{sv:g}）低于隔级上级打分（{sk:g}）"
                    )

    return [f"{display_name}：{msg}" for msg in issues]


def check_file(path: Path) -> tuple[list[str], bool]:
    """返回 (问题行列表, 是否识别为有效绩效表)。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return [f"(无法打开文件 {path}: {e})"], True

    all_issues: list[str] = []
    name = ""
    for ws in wb.worksheets:
        n = sheet_name_from_c2(ws)
        if n:
            name = n
        label = n or ws.title or path.name
        all_issues.extend(check_sheet(ws, label))

    if not name:
        # 无法从 C2 取姓名时，打印仍用工作表名或文件名
        pass
    return all_issues, True


def modify_sheet(ws) -> int:
    """修正当前表，返回写入的单元格数量。"""
    n = 0

    bounds = task_section_bounds(ws)
    if bounds:
        first_r, last_r = bounds
        for r in range(first_r, last_r + 1):
            if not ws.cell(row=r, column=4).value:
                continue
            st = to_float(ws.cell(row=r, column=COL_SELF_TIME).value)
            sq = to_float(ws.cell(row=r, column=COL_SELF_QUALITY).value)
            lt = to_float(ws.cell(row=r, column=COL_LEADER_TIME).value)
            lq = to_float(ws.cell(row=r, column=COL_LEADER_QUALITY_DIRECT).value)
            mq = to_float(ws.cell(row=r, column=COL_LEADER_QUALITY_SKIP).value)

            if st is not None and lt is not None and st < lt:
                ws.cell(row=r, column=COL_LEADER_TIME).value = st
                n += 1
            if sq is not None:
                if lq is not None and sq < lq:
                    ws.cell(row=r, column=COL_LEADER_QUALITY_DIRECT).value = sq
                    n += 1
                if mq is not None and sq < mq:
                    ws.cell(row=r, column=COL_LEADER_QUALITY_SKIP).value = sq
                    n += 1

    pb = personal_section_bounds(ws)
    if pb:
        p_first, p_last = pb
        for r in range(p_first, p_last + 1):
            sv = to_float(ws.cell(row=r, column=COL_PERSONAL_SELF).value)
            md = to_float(ws.cell(row=r, column=COL_PERSONAL_DIRECT).value)
            sk = to_float(ws.cell(row=r, column=COL_PERSONAL_SKIP).value)

            if sv is not None:
                if md is not None and sv < md:
                    ws.cell(row=r, column=COL_PERSONAL_DIRECT).value = sv
                    n += 1
                if sk is not None and sv < sk:
                    ws.cell(row=r, column=COL_PERSONAL_SKIP).value = sv
                    n += 1

    return n


def modify_file(path: Path) -> tuple[int, str]:
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return 0, f"无法打开: {e}"

    total = 0
    try:
        for ws in wb.worksheets:
            total += modify_sheet(ws)
        wb.save(path)
    except Exception as e:
        return 0, f"保存失败: {e}"
    return total, ""


def cmd_check(directory: str, result_path: Path) -> int:
    paths = iter_xlsx_files(directory)
    if not paths:
        print("未找到符合条件的 .xlsx 文件", file=sys.stderr)
        return 1

    bad_files: list[Path] = []
    exit_code = 0

    for path in paths:
        issues, _ok = check_file(path)
        if issues:
            # 无法打开等也算「有问题」
            bad_files.append(path)
            exit_code = 1
            for line in issues:
                print(line)

    result_path.write_text(
        "\n".join(str(p) for p in bad_files) + ("\n" if bad_files else ""),
        encoding="utf-8",
    )
    print(f"\n共 {len(bad_files)} 个文件写入 {result_path}", file=sys.stderr)
    return exit_code


def cmd_modify(result_path: Path) -> int:
    if not result_path.is_file():
        print(f"找不到结果文件: {result_path}", file=sys.stderr)
        return 1

    text = result_path.read_text(encoding="utf-8")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        print("结果文件为空，无需修改", file=sys.stderr)
        return 0

    code = 0
    for line in lines:
        p = Path(line)
        if not p.is_file():
            print(f"跳过（文件不存在）: {line}", file=sys.stderr)
            code = 1
            continue
        n, err = modify_file(p)
        if err:
            print(f"{p}: {err}", file=sys.stderr)
            code = 1
            continue
        print(f"{p}: 已调整 {n} 个单元格")
    return code


def main() -> None:
    parser = argparse.ArgumentParser(description="检查/修正自评分与上级分不一致")
    sub = parser.add_subparsers(dest="command", required=True)

    p_check = sub.add_parser("check", help="检查目录内所有 xlsx")
    p_check.add_argument("directory", help="要遍历的根目录")
    p_check.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help=f"结果文件路径（默认：当前目录/{RESULT_FILENAME}）",
    )

    p_mod = sub.add_parser("modify", help="按结果文件修正表格")
    p_mod.add_argument(
        "-i",
        "--input",
        type=Path,
        default=None,
        help=f"结果文件路径（默认：当前目录/{RESULT_FILENAME}）",
    )

    args = parser.parse_args()
    if args.command == "check":
        out = args.output if args.output is not None else Path.cwd() / RESULT_FILENAME
        raise SystemExit(cmd_check(args.directory, out))
    if args.command == "modify":
        inp = args.input if args.input is not None else Path.cwd() / RESULT_FILENAME
        raise SystemExit(cmd_modify(inp))


if __name__ == "__main__":
    main()
