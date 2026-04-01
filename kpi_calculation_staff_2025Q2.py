import openpyxl
import os
import json
from pathlib import Path



def check_format_and_calculate_performance(file_path, deduction_json):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 格式检查功能
    try:
        # 检查C列第二行是否填写姓名
        name = sheet['C2'].value
        if not name:
            raise ValueError("员工未填写姓名")

        # 检查是否存在“业务完成综合分（项目平均分）”行
        end_row = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "业务完成综合分（项目平均分）":
                end_row = row - 1
                break
        if end_row is None:
            raise ValueError("绩效目标截止行缺失")

        # 检查D7单元格和F8单元格
        if sheet['D7'].value != "任务目标" or sheet['F8'].value != "底线值":
            raise ValueError("绩效目标起始行错误")

    except ValueError as e:
        print(file_path)
        print(e)
        return

    # 绩效计算功能
    start_row = 9
    task_scores = []
    weights = []
    time_scores = []
    quality_scores = []

    for row in range(start_row, end_row + 1):
        task_name = sheet.cell(row=row, column=4).value  # D列
        weight = sheet.cell(row=row, column=5).value  # E列
        time_score = sheet.cell(row=row, column=12).value  # L列
        quality_score1 = sheet.cell(row=row, column=13).value  # M列
        quality_score2 = sheet.cell(row=row, column=14).value  # N列

        if not task_name:
            continue

        if weight is None:
            raise ValueError(f"{name}任务权重未填写")

        if quality_score1 is None:
            raise ValueError(f"{name}第{row}行未填写绩效分")

        if quality_score2 is None:
            quality_score2 = quality_score1

        quality_score = (quality_score1 + quality_score2) / 2
        task_scores.append((time_score * 0.5 + quality_score * 0.5) * weight)
        time_scores.append(time_score * weight)
        quality_scores.append(quality_score * weight)
        weights.append(weight)

    # 检查权重和是否为1，允许一个小的误差范围
    if abs(sum(weights) - 1) > 1e-6:
        print(weights)
        raise ValueError(f"{name}任务权重不为1")

    final_time_score = sum(time_scores)
    final_quality_score = sum(quality_scores)
    task_score = sum(task_scores)
    overtime_penalty = final_time_score < 90 or final_quality_score < 80

    # 综合素质分计算
    personal_scores = []
    start_personal = None
    end_personal = None

    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "创新能力":
            start_personal = row
        if sheet.cell(row=row, column=1).value == "遵章守纪":
            end_personal = row
            break

    if start_personal and end_personal:
        for row in range(start_personal, end_personal + 1):
            weight = sheet.cell(row=row, column=12).value  # L列
            score1 = sheet.cell(row=row, column=14).value  # N列

            score2 = sheet.cell(row=row, column=15).value  # O列

            if score2 is None:
                score2 = score1

            if isinstance(score1, str):
                score1 = int(score1)
            if isinstance(score2, str):
                score2 = int(score2)
            personal_scores.append(((score1 + score2) / 2) * weight)

    personal_score = sum(personal_scores)
    final_score = task_score * 0.7 + personal_score * 0.3

    # 输出结果
    # output = (f"{name}： 总完成时间分：{final_time_score:.2f}，总完成质量分：{final_quality_score:.2f}，"
    #           f"总项目分：{task_score:.2f}，综合素质分：{personal_score:.2f}，"
    #           f"最终绩效分：{final_score:.2f}，勤勉度扣分：{overtime_penalty}")
    # print(output)


    deduction = json.loads(deduction_json.read_text(encoding="utf-8")).get(name, {}).get("deduction", -1)
    return Member(name, final_time_score, final_quality_score, task_score, personal_score, final_score, overtime_penalty, deduction)

class Member:
    def __init__(self, name, final_time_score, final_quality_score, task_score, personal_score, final_score, overtime_penalty, deduction):
        self.name = name
        self.final_time_score = final_time_score
        self.final_quality_score = final_quality_score
        self.task_score = task_score
        self.personal_score = personal_score
        self.final_score = final_score
        self.overtime_penalty = overtime_penalty
        if self.overtime_penalty:
            self.deduction_ = deduction
        else:
            self.deduction_ = 0.0
        self.final_score_include_deduction = self.final_score - self.deduction_

    def __repr__(self):
        return (f"Member(name={self.name}, final_time_score={self.final_time_score:.2f}, "
                f"final_quality_score={self.final_quality_score:.2f}, task_score={self.task_score:.2f}, "
                f"personal_score={self.personal_score:.2f}, final_score={self.final_score:.2f}, "
                f"overtime_penalty={self.overtime_penalty})")

    def __lt__(self, other):
        if not isinstance(other, Member):
            return NotImplemented
        return self.final_score_include_deduction < other.final_score_include_deduction

    def __eq__(self, other):
        if not isinstance(other, Member):
            return NotImplemented
        return self.final_score_include_deduction == other.final_score_include_deduction

    def show(self):
        output = (f"{self.name}： \t总完成时间分：{self.final_time_score:.2f}，总完成质量分：{self.final_quality_score:.2f}，"
                  f"总项目分：{self.task_score:.2f}，综合素质分：{self.personal_score:.2f}，"
                  f"最终绩效分：{self.final_score:.2f}，勤勉度扣分：{self.overtime_penalty}, 考虑扣分的最终得分：{self.final_score_include_deduction:.2f}")
        print(output)

# 调用函数
# check_format_and_calculate_performance('test.xlsx')




def find_xlsx_files(directory):
    xlsx_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                xlsx_files.append(os.path.join(root, file))
    return xlsx_files

# 示例用法
directory = '/home/zhenghao/Program/kpi/data/2025Q4-staff/2025Q4-staff'  # 替换为你的文件夹路径
deduction_json = Path('/home/zhenghao/Program/kpi/data/2025Q4-staff/deduction.json')
xlsx_files = find_xlsx_files(directory)
member_list = []
for file_path in xlsx_files:
    member = check_format_and_calculate_performance(file_path, deduction_json)
    member_list.append(member)
member_list.sort(reverse=True)
for member in member_list:
    member.show()