#!/usr/bin/env python3
"""
遍历目录下所有 .xlsx（排除隐藏文件与 Office/WPS 锁文件），在绩效表中按规则填写「隔级上级」分数：

规则：隔级上级分 = floor(直接上级分 / 5) * 5 - 5（向下按 5 分一档取整后再减 5）。
例：93 -> 90 -> 85。

仅写入单元格的数值，不改动填充色等样式（openpyxl 对已有单元格只改 value 时通常保留样式）。

表结构约定（与「研发中心季度绩效考评表-专业线」样例一致）：
- 任务目标区：第 8 行表头中「直接上级」「隔级上级」相邻；数据行在「业务完成综合分（项目平均分）」之上，且 D 列有任务描述。
- 综合素质区：A 列为「评估维度」的行为表头行；数据行为「创新能力」至「遵章守纪」或「工作计划」。

用法：填写并保存 —— ``python fill_skip_level_scores.py <目录>``；仅校验不写文件 —— 加 ``--verify``。
"""

from __future__ import annotations

import argparse
import math
import os
from pathlib import Path

import openpyxl


def iter_xlsx_files(directory: str) -> list[Path]:
    root = Path(directory).resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"不是目录: {root}")
    out: list[Path] = []
    for dirpath, _dirnames, filenames in os.walk(root):
        for name in filenames:
            if name.startswith("."):
                continue
            if not name.endswith(".xlsx"):
                continue
            if name.startswith("~$") or name.startswith(".~"):
                continue
            out.append(Path(dirpath) / name)
    return sorted(out)


def _to_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def skip_level_from_direct(direct) -> int | None:
    x = _to_float(direct)
    if x is None:
        return None
    return int(math.floor(x / 5.0) * 5.0 - 5.0)


def _find_adjacent_direct_skip_columns(ws, row: int) -> tuple[int, int] | None:
    direct_col: int | None = None
    skip_col: int | None = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v == "直接上级":
            direct_col = c
        elif v == "隔级上级":
            skip_col = c
    if direct_col is None or skip_col is None:
        return None
    if skip_col != direct_col + 1:
        return None
    return direct_col, skip_col


def _task_header_direct_skip(ws) -> tuple[int, int, int] | None:
    """任务目标区：在若干行内查找相邻的「直接上级」「隔级上级」列。返回 (header_row, direct_col, skip_col)。"""
    for r in range(7, min(11, ws.max_row + 1)):
        pair = _find_adjacent_direct_skip_columns(ws, r)
        if pair:
            return r, pair[0], pair[1]
    return None


def _task_section_bounds(ws) -> tuple[int, int] | None:
    """返回 (first_data_row, last_data_row)，找不到则 None。"""
    if ws["D7"].value != "任务目标":
        return None
    end_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v == "业务完成综合分（项目平均分）":
            end_row = r - 1
            break
    if end_row is None or end_row < 9:
        return None
    return 9, end_row


def _personal_bounds(ws) -> tuple[int, int, int] | None:
    """返回 (first_data_row, last_data_row, 评估维度表头行)。"""
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "评估维度":
            header_row = r
            break
    if header_row is None or _find_adjacent_direct_skip_columns(ws, header_row) is None:
        return None

    start_row = None
    end_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a == "创新能力":
            start_row = r
            break
    if start_row is None:
        return None

    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a == "遵章守纪":
            end_row = r
            break
        if a == "工作计划":
            end_row = r
            break
    if end_row is None:
        return None
    return start_row, end_row, header_row


def fill_sheet(ws) -> tuple[int, int]:
    """
    填写当前工作表。返回 (任务区写入格数, 综合素质区写入格数)。
    """
    task_filled = 0
    personal_filled = 0

    bounds = _task_section_bounds(ws)
    if bounds:
        first_r, last_r = bounds
        t_head = _task_header_direct_skip(ws)
        if t_head:
            _hr, direct_col, skip_col = t_head
            for r in range(first_r, last_r + 1):
                if not ws.cell(row=r, column=4).value:
                    continue
                direct = ws.cell(row=r, column=direct_col).value
                s = skip_level_from_direct(direct)
                if s is None:
                    continue
                cell = ws.cell(row=r, column=skip_col)
                cell.value = s
                task_filled += 1

    pb = _personal_bounds(ws)
    if pb:
        p_first, p_last, header_row = pb
        pair = _find_adjacent_direct_skip_columns(ws, header_row)
        if pair:
            direct_col, skip_col = pair
            for r in range(p_first, p_last + 1):
                direct = ws.cell(row=r, column=direct_col).value
                s = skip_level_from_direct(direct)
                if s is None:
                    continue
                cell = ws.cell(row=r, column=skip_col)
                cell.value = s
                personal_filled += 1

    return task_filled, personal_filled


def _actual_matches_expected(actual, expected: int) -> bool:
    ax = _to_float(actual)
    if ax is None:
        return False
    return int(ax) == int(expected)


def verify_sheet(ws, sheet_label: str = "") -> list[str]:
    """
    校验已有「隔级上级」是否与直接上级按规则一致。
    仅在有数值型「直接上级」时检查；直接上级为空则不报错。
    """
    errs: list[str] = []
    pref = f"[{sheet_label}] " if sheet_label else ""

    bounds = _task_section_bounds(ws)
    if bounds:
        first_r, last_r = bounds
        t_head = _task_header_direct_skip(ws)
        if t_head:
            _hr, direct_col, skip_col = t_head
            for r in range(first_r, last_r + 1):
                if not ws.cell(row=r, column=4).value:
                    continue
                direct = ws.cell(row=r, column=direct_col).value
                exp = skip_level_from_direct(direct)
                if exp is None:
                    continue
                act = ws.cell(row=r, column=skip_col).value
                if not _actual_matches_expected(act, exp):
                    errs.append(
                        f"{pref}任务区 行{r}: 直接上级={direct!r} 时期望隔级={exp} 实际={act!r}"
                    )

    pb = _personal_bounds(ws)
    if pb:
        p_first, p_last, header_row = pb
        pair = _find_adjacent_direct_skip_columns(ws, header_row)
        if pair:
            direct_col, skip_col = pair
            for r in range(p_first, p_last + 1):
                direct = ws.cell(row=r, column=direct_col).value
                exp = skip_level_from_direct(direct)
                if exp is None:
                    continue
                act = ws.cell(row=r, column=skip_col).value
                if not _actual_matches_expected(act, exp):
                    errs.append(
                        f"{pref}综合素质 行{r}: 直接上级={direct!r} 时期望隔级={exp} 实际={act!r}"
                    )

    return errs


def verify_file(path: Path) -> list[str]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return [f"无法打开: {e}"]
    all_e: list[str] = []
    for ws in wb.worksheets:
        all_e.extend(verify_sheet(ws, ws.title))
    return all_e


def process_file(path: Path) -> tuple[int, int, str]:
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return 0, 0, f"无法打开: {e}"

    total_task = total_personal = 0
    try:
        for sheet in wb.worksheets:
            t, p = fill_sheet(sheet)
            total_task += t
            total_personal += p
        wb.save(path)
    except Exception as e:
        return 0, 0, f"处理失败: {e}"

    return total_task, total_personal, ""


def main() -> None:
    parser = argparse.ArgumentParser(description="按直接上级分填写隔级上级分（目录内所有 xlsx）")
    parser.add_argument(
        "directory",
        help="要遍历的根目录（会递归子目录）",
    )
    parser.add_argument(
        "--verify",
        action="store_true",
        help="不写入，仅校验当前文件中「隔级上级」是否与规则一致（需已保存的数值）",
    )
    args = parser.parse_args()
    paths = iter_xlsx_files(args.directory)
    if not paths:
        print("未找到符合条件的 .xlsx 文件")
        return

    if args.verify:
        bad = 0
        for path in paths:
            errs = verify_file(path)
            if errs:
                bad += 1
                print(f"{path}")
                for e in errs:
                    print(f"  {e}")
            else:
                print(f"{path}: 校验通过")
        if bad:
            raise SystemExit(1)
        return

    for path in paths:
        res = process_file(path)
        if res[2]:
            print(f"{path}: {res[2]}")
            continue
        tt, tp, _ = res
        print(f"{path}: 任务区 {tt} 格, 综合素质 {tp} 格")


if __name__ == "__main__":
    main()
